import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell import Cell
from xlrd import sheet
from openpyxl import Workbook
from openpyxl import worksheet as ws


data = pd.read_excel(r'C:\Users\User\Desktop\pyth1\week.xlsx')

df = pd.DataFrame(data, columns=['Luni', 'Marti', 'Miercuri', 'Joi', 'Vineri', 'Sambata/Duminica'])

wb = load_workbook(filename='week.xlsx')
sheet_ranges = wb['Sheet2']

print(sheet_ranges.merged_cells.ranges)
df2=df

def testMerge(row, column):
  cell=sheet_ranges.cell(row, column)
  for mergedCell in sheet_ranges.merged_cells:
    if cell.coordinate in mergedCell:
      return True,mergedCell.start_cell.value
  return False,0

for index, row in df.iterrows():
  if index==0:
    continue
  for i in range(len(df.columns)):
    result,val=testMerge(index+1, i+1)
    print(result,val)
    if result:
      df2.iloc[[index-1],[i]]= val


''' 
for index, row in df.iterrows():
  print(row['Vineri'])
'''

#n=testMerge(2,5)
#n2=testMerge(5,5)

df2 = df2.style.set_properties(**{'font-weight': 'bold'})

df2.to_excel(r'C:\Users\User\Desktop\pyth1\week2.xlsx', index=False, header=True)


